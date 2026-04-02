"""Test core modules of the PC attendance app."""
import sys
import os
import sqlite3
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def create_test_db():
    test_db = os.path.join(tempfile.gettempdir(), 'test_attendance.db')
    if os.path.exists(test_db):
        os.remove(test_db)
    conn = sqlite3.connect(test_db)
    c = conn.cursor()
    c.executescript("""
CREATE TABLE IF NOT EXISTS departments (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, created_at TEXT DEFAULT '', updated_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, created_at TEXT DEFAULT '', updated_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL, emp_id TEXT NOT NULL UNIQUE, name TEXT NOT NULL, father_name TEXT DEFAULT '', department_id INTEGER, category_id INTEGER, dob TEXT DEFAULT '', doj TEXT DEFAULT '', weekly_off_day INTEGER DEFAULT 0, fd TEXT DEFAULT '', ot_rr_type TEXT DEFAULT 'COMP', ot_rr_value REAL DEFAULT 0, gross_salary REAL DEFAULT 0, is_active INTEGER DEFAULT 1, created_at TEXT DEFAULT '', updated_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS attendance_records (id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER NOT NULL, date TEXT NOT NULL, in_time TEXT DEFAULT '', out_time TEXT DEFAULT '', status TEXT DEFAULT 'P', total_hours REAL DEFAULT 0, ot_hours REAL DEFAULT 0, remarks TEXT DEFAULT '', created_at TEXT DEFAULT '', updated_at TEXT DEFAULT '', UNIQUE(employee_id, date));
CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL, updated_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS unit_locations (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, unit_number TEXT DEFAULT '', address TEXT DEFAULT '', is_active INTEGER DEFAULT 1, created_at TEXT DEFAULT '', updated_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS punch_log (id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER NOT NULL, date TEXT NOT NULL, time TEXT NOT NULL, punch_type TEXT NOT NULL, unit_id INTEGER, recorded_by TEXT DEFAULT 'supervisor', created_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS backup_log (id INTEGER PRIMARY KEY AUTOINCREMENT, backup_type TEXT NOT NULL, file_path TEXT DEFAULT '', status TEXT NOT NULL, error_message TEXT DEFAULT '', created_at TEXT DEFAULT '');

INSERT INTO settings (key, value) VALUES ('normal_work_hours', '8.0');
INSERT INTO settings (key, value) VALUES ('unit_number', '125');
INSERT INTO settings (key, value) VALUES ('company_name', 'Test Factory');

INSERT INTO departments (id, name) VALUES (1, 'Production');
INSERT INTO departments (id, name) VALUES (2, 'Accounts');
INSERT INTO departments (id, name) VALUES (3, 'Maintenance');

INSERT INTO categories (id, name) VALUES (1, 'Staff');
INSERT INTO categories (id, name) VALUES (2, 'Worker');

INSERT INTO employees (code, emp_id, name, father_name, department_id, category_id, dob, doj, weekly_off_day, fd, ot_rr_type, ot_rr_value, gross_salary) VALUES ('X2169', 'E0347', 'NARINDER SINGH THAKUR', 'AJIT SINGH THAKUR', 1, 2, '', '2021-09-15', 0, '', 'COMP', 0, 281533);
INSERT INTO employees (code, emp_id, name, father_name, department_id, category_id, dob, doj, weekly_off_day, ot_rr_type, ot_rr_value, gross_salary) VALUES ('X2426', 'E0379', 'RAHUL SINGHAL', 'LT MR. SINGHAL', 2, 1, '', '', 0, 'COMP', 0, 155000);
INSERT INTO employees (code, emp_id, name, father_name, department_id, category_id, dob, doj, weekly_off_day, fd, ot_rr_type, ot_rr_value, gross_salary) VALUES ('1149', 'E0244', 'SANDEEP SINGH', 'SHEELENDRA SINGH', 2, 2, '', '2016-01-01', 0, 'D', 'RR', 45, 27300);

INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (1, '2026-02-01', '', '', 'W', 0, 8.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (1, '2026-02-02', '09:00', '17:00', 'P', 8.0, 0.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (1, '2026-02-03', '09:00', '19:00', 'P', 10.0, 2.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (2, '2026-02-01', '', '', 'W', 0, 0.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (2, '2026-02-02', '09:00', '17:00', 'P', 8.0, 0.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (3, '2026-02-01', '', '', 'W', 0, 0.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (3, '2026-02-02', '09:00', '18:00', 'P', 9.0, 1.0);
INSERT INTO attendance_records (employee_id, date, in_time, out_time, status, total_hours, ot_hours) VALUES (3, '2026-02-03', '09:00', '20:00', 'P', 11.0, 3.0);
    """)
    conn.commit()
    conn.close()
    return test_db


def main():
    # Test imports
    print("Testing imports...")
    from src.database.connection import DatabaseManager
    print("  DatabaseManager OK")
    from src.database.models import Department, Category, Employee, AttendanceRecord, Setting
    print("  Models OK")
    from src.database import queries
    print("  Queries OK")
    from src.util.date_utils import get_days_in_month, get_day_of_week, get_month_name, format_date_display, get_week_day_index
    print("  DateUtils OK")
    from src.export.column_config import ColumnDefinition, get_default_columns
    print("  ColumnConfig OK")

    # Test date utils
    print("\nTesting date_utils...")
    assert get_days_in_month(2026, 2) == 28, f"Expected 28, got {get_days_in_month(2026, 2)}"
    print(f"  Days in Feb 2026: {get_days_in_month(2026, 2)}")
    print(f"  Day of week Mar 30, 2026: {get_day_of_week(2026, 3, 30)}")
    print(f"  Month name 2: {get_month_name(2)}")
    print(f"  Format date: {format_date_display('2021-09-15')}")
    print(f"  Week day index (Mar 29=Sun): {get_week_day_index(2026, 3, 29)}")

    # Test column config
    print("\nTesting column_config...")
    cols = get_default_columns()
    print(f"  Default columns: {len(cols)}")
    for c in cols:
        print(f"    {c.key}: {c.label} (width={c.width})")

    # Create and load test DB
    print("\nTesting DatabaseManager...")
    test_db = create_test_db()
    db = DatabaseManager()
    assert not db.is_loaded()
    db.load_database(test_db)
    assert db.is_loaded()
    print(f"  Loaded: {db.is_loaded()}, Path: {db.get_path()}")

    # Test queries
    print("\nTesting queries...")
    depts = queries.get_all_departments(db)
    print(f"  Departments: {[d.name for d in depts]}")
    assert len(depts) == 3

    cats = queries.get_all_categories(db)
    print(f"  Categories: {[c.name for c in cats]}")
    assert len(cats) == 2

    emps = queries.get_all_employees(db)
    print(f"  Employees: {len(emps)}")
    assert len(emps) == 3
    for e in emps:
        print(f"    {e.code} {e.emp_id} {e.name} dept={e.department_name} cat={e.category_name} salary={e.gross_salary}")

    settings = queries.get_settings(db)
    print(f"  Settings: {settings}")
    assert settings['normal_work_hours'] == '8.0'
    assert settings['unit_number'] == '125'

    att = queries.get_attendance_for_month(db, 2026, 2)
    print(f"  Feb 2026 attendance records: {len(att)}")
    assert len(att) == 8

    report = queries.get_monthly_report_data(db, 2026, 2)
    print(f"  Monthly report rows: {len(report)}")
    assert len(report) == 3
    for r in report:
        print(f"    {r['code']} {r['name']}: days={r['days']}")

    # Test Excel export
    print("\nTesting Excel export...")
    from src.export.excel_exporter import ExcelExporter
    print("  ExcelExporter imported OK")
    exporter = ExcelExporter(db)
    output_path = os.path.join(tempfile.gettempdir(), 'test_attendance_export.xlsx')
    exporter.export(2026, 2, '125', get_default_columns(), output_path)
    assert os.path.exists(output_path), "Excel file not created"
    file_size = os.path.getsize(output_path)
    print(f"  Exported to: {output_path}")
    print(f"  File size: {file_size} bytes")

    # Verify Excel content
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    print(f"  Sheet name: {ws.title}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"  Title (A1): {ws.cell(1, 1).value}")
    wb.close()

    # Cleanup
    db.close()
    os.remove(test_db)
    os.remove(output_path)

    print("\n=== ALL TESTS PASSED ===")


if __name__ == "__main__":
    main()
