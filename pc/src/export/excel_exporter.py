import calendar
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.database.connection import DatabaseManager
from src.database.queries import get_monthly_report_data
from src.export.column_config import ColumnDefinition
from src.util.date_utils import get_day_of_week, get_days_in_month, get_month_name

# Weekly off code mapping: int index -> abbreviation
WEEKDAY_NAMES = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

# Status fill colours
FILL_ABSENT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_WEEKLY_OFF = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_COMP_OFF = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, color="C00000", size=14)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


class ExcelExporter:
    """Exports monthly attendance data to an Excel workbook."""

    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    def export(
        self,
        year: int,
        month: int,
        unit_number: str,
        columns: List[ColumnDefinition],
        output_path: str,
    ) -> None:
        report_data = get_monthly_report_data(self.db_manager, year, month)
        days_in_month = get_days_in_month(year, month)
        month_name = get_month_name(month)

        num_emp_cols = len(columns)
        # Each day occupies 2 columns: status + OT
        total_cols = num_emp_cols + days_in_month * 2

        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name}-{year}"

        # ===== Row 1: Title (merged across all columns) =====
        title_text = (
            f"ATTENDANCE SHEET FOR THE MONTH OF {month_name}-{year}- UNIT-({unit_number})"
        )
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=total_cols,
        )
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # ===== Row 2: Date numbers above day pairs =====
        # Employee columns are blank in row 2; day columns show date number merged across 2 cells
        for day in range(1, days_in_month + 1):
            status_col = num_emp_cols + (day - 1) * 2 + 1
            ot_col = status_col + 1
            ws.merge_cells(
                start_row=2, start_column=status_col,
                end_row=2, end_column=ot_col,
            )
            cell = ws.cell(row=2, column=status_col, value=day)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            ws.cell(row=2, column=ot_col).border = THIN_BORDER

        # ===== Row 3: "OT" sub-headers under each date =====
        for day in range(1, days_in_month + 1):
            status_col = num_emp_cols + (day - 1) * 2 + 1
            ot_col = status_col + 1
            # Status sub-column left blank (or can put day abbreviation later in row 4)
            ot_cell = ws.cell(row=3, column=ot_col, value="OT")
            ot_cell.font = HEADER_FONT
            ot_cell.alignment = CENTER_ALIGN
            ot_cell.border = THIN_BORDER
            ws.cell(row=3, column=status_col).border = THIN_BORDER

        # ===== Row 4: Column headers =====
        # Employee column headers
        for ci, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=ci, value=col_def.label)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Day-of-week abbreviations for each date
        for day in range(1, days_in_month + 1):
            day_abbr = get_day_of_week(year, month, day)
            status_col = num_emp_cols + (day - 1) * 2 + 1
            ot_col = status_col + 1

            cell = ws.cell(row=4, column=status_col, value=day_abbr)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

            ot_cell = ws.cell(row=4, column=ot_col)
            ot_cell.border = THIN_BORDER

        # ===== Data rows (row 5+) =====
        for row_idx, emp_data in enumerate(report_data):
            excel_row = 5 + row_idx
            serial = row_idx + 1

            for ci, col_def in enumerate(columns, start=1):
                value = self._get_employee_field(emp_data, col_def, serial)
                cell = ws.cell(row=excel_row, column=ci, value=value)
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

            days: Dict[int, Dict] = emp_data["days"]
            for day in range(1, days_in_month + 1):
                status_col = num_emp_cols + (day - 1) * 2 + 1
                ot_col = status_col + 1

                day_data = days.get(day, {"status": "", "ot_hours": 0.0})
                status = day_data["status"]
                ot_hours = day_data["ot_hours"]

                # Status cell
                s_cell = ws.cell(row=excel_row, column=status_col, value=status)
                s_cell.alignment = CENTER_ALIGN
                s_cell.border = THIN_BORDER
                if status == "A":
                    s_cell.fill = FILL_ABSENT
                elif status == "W":
                    s_cell.fill = FILL_WEEKLY_OFF
                elif status == "CO":
                    s_cell.fill = FILL_COMP_OFF

                # OT cell
                ot_cell = ws.cell(row=excel_row, column=ot_col, value=ot_hours)
                ot_cell.number_format = "0.0"
                ot_cell.alignment = CENTER_ALIGN
                ot_cell.border = THIN_BORDER

        # ===== Column widths =====
        for ci, col_def in enumerate(columns, start=1):
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = col_def.width

        for day in range(1, days_in_month + 1):
            status_col = num_emp_cols + (day - 1) * 2 + 1
            ot_col = status_col + 1
            ws.column_dimensions[get_column_letter(status_col)].width = 5
            ws.column_dimensions[get_column_letter(ot_col)].width = 5

        # ===== Freeze panes =====
        # Freeze at the top-left corner of the first day column, at row 5 (first data row)
        freeze_col_letter = get_column_letter(num_emp_cols + 1)
        ws.freeze_panes = f"{freeze_col_letter}5"

        wb.save(output_path)

    @staticmethod
    def _get_employee_field(emp_data: dict, col_def: ColumnDefinition, serial: int):
        """Extract the value for a given column definition from employee data."""
        key = col_def.field

        if key == "sr":
            return serial

        if key == "weekly_off_day":
            raw = emp_data.get("weekly_off_day", "SUN")
            # The model stores abbreviated codes like "SUN", "MON", etc.
            return raw

        if key == "ot_rr":
            ot_type = emp_data.get("ot_rr_type", "")
            ot_val = emp_data.get("ot_rr_value", 0.0)
            if ot_type:
                return f"{ot_type} {ot_val}"
            return ""

        if key == "gross_salary":
            return emp_data.get("gross_salary", 0.0)

        return emp_data.get(key, "")
